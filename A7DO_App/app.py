import math
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
import plotly.graph_objects as go

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="A7DO Genesis Mind — v6 Workbook",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Assume repo layout:
#   repo_root/
#       app/app.py
#       excel_report/a7do-v6/A7DO_DNA_Master_v6.xlsx
APP_DIR = Path(__file__).parent
REPO_ROOT = APP_DIR.parent
XLSX_PATH = REPO_ROOT / "excel_report" / "a7do-v6" / "A7DO_DNA_Master_v6.xlsx"


# ──────────────────────────────────────────────────────────────────────────────
# WORKBOOK HELPERS
# ──────────────────────────────────────────────────────────────────────────────

@st.cache_resource
def load_workbook():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    return wb


@st.cache_data
def sheet_to_df(sheet_name: str) -> pd.DataFrame:
    wb = load_workbook()
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            rows.append(list(row))
    if not rows:
        return pd.DataFrame()
    # First row as header if it looks like text
    header = rows[0]
    if all(isinstance(c, str) or c is None for c in header):
        return pd.DataFrame(rows[1:], columns=header)
    return pd.DataFrame(rows)


@st.cache_data
def get_sheet_names():
    return load_workbook().sheetnames


# ──────────────────────────────────────────────────────────────────────────────
# ORGANISM STATE (SURROGATE, MATCHING YOUR SPEC)
# ──────────────────────────────────────────────────────────────────────────────

def organism_state_from_tick(tick: int) -> dict:
    """Rebuilds the v6 surrogate organism state from Tick, matching the workbook spec."""
    week = round(tick / 80)

    # Height / Mass / HR from your logistic approximations
    height = 50 if week < 40 else min(50 + (177 - 50) * (1 - math.exp(-0.005 * (week - 40))), 177)
    mass = 3.5 if week < 40 else min(3.5 + (70 - 3.5) * (1 - math.exp(-0.004 * (week - 40))), 70)
    hr = 140 if week < 40 else max(70, 140 - (140 - 70) * ((week - 40) / 1160))

    vocab = round(50000 / (1 + math.exp(-0.05 * (week - 156)))) if week > 0 else 0

    motor = 5 if week >= 400 else (4 if week >= 260 else (3 if week >= 160 else (2 if week >= 80 else 1)))
    tom = 5 if week >= 624 else (4 if week >= 312 else (3 if week >= 208 else (2 if week >= 156 else 1)))
    perm = 3 if week >= 52 else (2 if week >= 44 else (1 if week >= 36 else 0))

    birth = tick >= 3200
    phase7 = tick >= 96000
    wisdom = min(0.1 + ((week - 1200) / 800) * 0.9, 1.0) if week >= 1200 else 0.0

    # Life stage map from DNA Loop Engine
    if week >= 1400:
        stage = "Elder"
    elif week >= 1200:
        stage = "Mature Adult"
    elif week >= 1100:
        stage = "Mid Adult"
    elif week >= 1000:
        stage = "Adult"
    elif week >= 936:
        stage = "Young Adult"
    elif week >= 624:
        stage = "Adolescent"
    elif week >= 260:
        stage = "Pre-Adolescent"
    elif week >= 156:
        stage = "Child"
    elif week >= 80:
        stage = "Toddler"
    elif week >= 52:
        stage = "Infant"
    elif week >= 40:
        stage = "Newborn"
    elif week >= 28:
        stage = "Fetal Late"
    elif week >= 12:
        stage = "Fetal Mid"
    elif week >= 4:
        stage = "Fetal Early"
    else:
        stage = "Embryo"

    if tick % 800 == 0:
        ll = "💤 Sleep Consolidation"
    elif tick % 10 == 0:
        ll = "🔁 Repetition"
    elif tick % 5 == 0:
        ll = "🤝 Interaction"
    else:
        ll = "👁️ Exposure"

    C = min(0.05 + week / 3000, 1.0)
    pred_err = max(0.1, math.exp(-0.0001 * tick))
    ltm = min(int(tick * 0.96), 200000)

    return dict(
        tick=tick,
        week=week,
        stage=stage,
        height=round(height, 1),
        mass=round(mass, 2),
        hr=round(hr, 1),
        vocab=vocab,
        motor=motor,
        tom=tom,
        perm=perm,
        birth=birth,
        phase7=phase7,
        wisdom=round(wisdom, 3),
        ll_phase=ll,
        C=round(C, 3),
        pred_err=round(pred_err, 3),
        ltm=ltm,
    )


# ──────────────────────────────────────────────────────────────────────────────
# PAGE RENDERERS
# ──────────────────────────────────────────────────────────────────────────────

def render_master_dashboard(state: dict):
    st.title("🧬 A7DO GENESIS MIND — Master Dashboard v6")
    st.caption(
        f"Tick {state['tick']:,} · Week {state['week']} · {state['stage']} · {state['ll_phase']}"
    )

    if state["phase7"]:
        st.success("🌟 Phase 7 ACTIVE — Wisdom / Creativity / Career / Legacy engines online")
    elif state["birth"]:
        st.info("🎉 Birth complete — postnatal development in progress")
    else:
        st.warning("⏳ Prenatal — Birth at Tick 3,200 (Week 40)")

    # Key metrics
    cols = st.columns(5)
    metrics = [
        ("📏 Height", f"{state['height']} cm"),
        ("⚖️ Mass", f"{state['mass']} kg"),
        ("❤️ Heart Rate", f"{state['hr']} bpm"),
        ("💬 Vocabulary", f"{state['vocab']:,} words"),
        ("🦾 Motor Stage", f"{state['motor']}/5"),
        ("🧠 ToM Stage", f"{state['tom']}/5"),
        ("👁️ Object Permanence", f"{state['perm']}/3"),
        ("✨ Consciousness C", f"{state['C']}"),
        ("💾 LTM Events", f"{state['ltm']:,}"),
        ("🦉 Wisdom W(t)", f"{state['wisdom']}"),
    ]
    for i, (label, val) in enumerate(metrics):
        with cols[i % 5]:
            st.metric(label, val)

    st.markdown("---")

    # Developmental phase timeline (from your spec)
    st.subheader("📈 Developmental Phase Timeline")
    phases = [
        ("Phase 1 — Biology & Embodiment", 0, 3200),
        ("Phase 2 — Sensorimotor", 3200, 6400),
        ("Phase 3 — Core Cognition", 6400, 12480),
        ("Phase 4 — Social Cognition", 12480, 49920),
        ("Phase 5 — Cultural & World", 49920, 74880),
        ("Phase 6 — Identity", 74880, 96000),
        ("Phase 7 — Wisdom", 96000, 160000),
    ]
    fig = go.Figure()
    for name, start, end in phases:
        fig.add_trace(
            go.Bar(
                x=[end - start],
                y=[name],
                base=[start],
                orientation="h",
                hovertemplate=f"{name}<br>Tick {start:,}–{end:,}<extra></extra>",
            )
        )
    fig.add_vline(
        x=state["tick"],
        line_color="#f97316",
        line_width=3,
        annotation_text=f"Tick {state['tick']:,}",
    )
    fig.update_layout(
        template="plotly_dark",
        height=320,
        showlegend=False,
        xaxis_title="Tick",
        margin=dict(l=0, r=0, t=20, b=0),
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    st.subheader("🗂️ Sheet Index (14 Sheets)")
    index_df = pd.DataFrame(
        [
            (0, "🏠 Master Dashboard", "Core", "Live organism state · phase timeline · sheet index"),
            (1, "🧬 DNA Loop Engine", "Core", "Master tick controller · lifecycle table · all derived formulas"),
            (2, "📐 Phase 1 — Biology & Embodiment", "Phase 1", "Prenatal phases · birth transition · body systems"),
            (3, "🦿 Phase 2 — Sensorimotor", "Phase 2", "Motor stages · rigid-body dynamics · reflexes"),
            (4, "🧠 Phase 3 — Core Cognition", "Phase 3", "Object permanence · episodic memory · value system"),
            (5, "🤝 Phase 4 — Social Cognition", "Phase 4", "Language grounding · Theory of Mind · seed vocabulary"),
            (6, "🌍 Phase 5 — Cultural & World", "Phase 5", "Cultural stages · NPC network · BeenFore City"),
            (7, "🪪 Phase 6 — Identity", "Phase 6", "Identity stages · equations · 22-dim skill graph"),
            (8, "✨ Phase 7 — Wisdom", "Phase 7", "EQ_CREAT_17 · EQ_WISDOM_18 · EQ_CAREER_19 · EQ_LEGACY_20"),
            (9, "🚀 Phase 8 — AGI Architecture", "Phase 8", "AGI readiness · Vision V1 · Motor M1 · Planning P2"),
            (10, "🔄 Learning Loop", "Learning", "4-stage cycle · word pipeline · web-hook learning"),
            (11, "⚙️ System Reference", "Meta", "52 parameters · 24 engine registry · tick schedule"),
            (12, "👩 Lorraine — Parent Profile", "Profile", "Full biological · psychological · social profile"),
            (13, "👨 China — Parent Profile", "Profile", "Full biological · psychological · social profile"),
        ],
        columns=["#", "Sheet", "Category", "Key Content"],
    )
    st.dataframe(index_df, use_container_width=True, hide_index=True)


def render_sheet_view(sheet_name: str):
    st.subheader(f"📄 Sheet: {sheet_name}")
    df = sheet_to_df(sheet_name)
    if df.empty:
        st.info("No data or sheet not found.")
    else:
        st.dataframe(df, use_container_width=True, hide_index=True)


def render_phase_page(title: str, sheet_name: str):
    st.title(title)
    st.caption("Source: A7DO_DNA_Master_v6 · Change Tick in sidebar to explore lifecycle.")
    render_sheet_view(sheet_name)


def render_learning_loop():
    st.title("🔄 Learning Loop — Experience-First Architecture")
    render_sheet_view("🔄 Learning Loop")


def render_system_reference():
    st.title("⚙️ System Reference — Parameters · Equations · Engine Registry")
    render_sheet_view("⚙️ System Reference")


def render_profiles():
    tab1, tab2 = st.tabs(["👩 Lorraine — Parent Profile", "👨 China — Parent Profile"])
    with tab1:
        render_sheet_view("👩 Lorraine — Parent Profile")
    with tab2:
        render_sheet_view("👨 China — Parent Profile")


def render_all_sheets_explorer():
    st.title("📊 All Sheets Explorer")
    names = get_sheet_names()
    sheet = st.selectbox("Select sheet", names)
    render_sheet_view(sheet)


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR & ROUTER
# ──────────────────────────────────────────────────────────────────────────────

if "tick" not in st.session_state:
    st.session_state.tick = 0

with st.sidebar:
    st.markdown("## 🧬 A7DO Genesis Mind — v6")
    st.markdown("Clean 14-sheet rebuild · All phases 1–8")
    st.divider()

    tick = st.slider("⚡ Current Tick", 0, 160000, st.session_state.tick, step=80)
    st.session_state.tick = tick
    state = organism_state_from_tick(tick)
    st.caption(f"Week {state['week']} · {state['stage']}")
    st.divider()

    pages = [
        "🏠 Master Dashboard",
        "📐 Phase 1 — Biology & Embodiment",
        "🦿 Phase 2 — Sensorimotor",
        "🧠 Phase 3 — Core Cognition",
        "🤝 Phase 4 — Social Cognition",
        "🌍 Phase 5 — Cultural & World",
        "🪪 Phase 6 — Identity",
        "✨ Phase 7 — Wisdom",
        "🚀 Phase 8 — AGI Architecture",
        "🔄 Learning Loop",
        "⚙️ System Reference",
        "👤 Parent Profiles",
        "📊 All Sheets Explorer",
    ]
    page = st.radio("Navigate", pages, label_visibility="collapsed")

# ──────────────────────────────────────────────────────────────────────────────
# ROUTE
# ──────────────────────────────────────────────────────────────────────────────

if page == "🏠 Master Dashboard":
    render_master_dashboard(state)

elif page == "📐 Phase 1 — Biology & Embodiment":
    render_phase_page("📐 Phase 1 — Biology & Embodiment", "📐 Phase 1 — Biology & Embodiment")

elif page == "🦿 Phase 2 — Sensorimotor":
    render_phase_page("🦿 Phase 2 — Sensorimotor", "🦿 Phase 2 — Sensorimotor")

elif page == "🧠 Phase 3 — Core Cognition":
    render_phase_page("🧠 Phase 3 — Core Cognition", "🧠 Phase 3 — Core Cognition")

elif page == "🤝 Phase 4 — Social Cognition":
    render_phase_page("🤝 Phase 4 — Social Cognition", "🤝 Phase 4 — Social Cognition")

elif page == "🌍 Phase 5 — Cultural & World":
    render_phase_page("🌍 Phase 5 — Cultural & World", "🌍 Phase 5 — Cultural & World")

elif page == "🪪 Phase 6 — Identity":
    render_phase_page("🪪 Phase 6 — Identity", "🪪 Phase 6 — Identity")

elif page == "✨ Phase 7 — Wisdom":
    render_phase_page("✨ Phase 7 — Wisdom", "✨ Phase 7 — Wisdom")

elif page == "🚀 Phase 8 — AGI Architecture":
    render_phase_page("🚀 Phase 8 — AGI Architecture", "🚀 Phase 8 — AGI Architecture")

elif page == "🔄 Learning Loop":
    render_learning_loop()

elif page == "⚙️ System Reference":
    render_system_reference()

elif page == "👤 Parent Profiles":
    render_profiles()

elif page == "📊 All Sheets Explorer":
    render_all_sheets_explorer()
