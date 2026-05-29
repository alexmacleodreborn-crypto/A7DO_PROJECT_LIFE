"""
A7DO Global Parameters — loaded from workbook ⚙️ Parameters sheet.
Falls back to canonical defaults if workbook unavailable.
"""
from dataclasses import dataclass
from pathlib import Path
import openpyxl


@dataclass
class A7DOParams:
    # ── EQ_DNA_01 ─────────────────────────────────────────────────────────────
    alpha: float = 0.05     # Growth coefficient (primary)
    beta: float  = 0.05     # Growth coefficient (secondary)
    gamma: float = 0.05     # Maturity index coefficient

    # ── EQ_ANAT_02 ────────────────────────────────────────────────────────────
    delta: float = 0.5      # Cell division rate constant
    epsilon: float = 0.5    # Nutrient flow constant
    zeta: float  = 0.5      # Oxygen exchange constant
    eta: float   = 0.5      # Waste removal rate

    # ── EQ_NEUR_03 ────────────────────────────────────────────────────────────
    theta: float = 0.025    # Synapse density coefficient
    iota: float  = 0.025    # Neural gain coefficient
    kappa: float = 0.025    # Input signal multiplier
    lam: float   = 0.025    # Noise attenuation factor

    # ── EQ_CIRC_04 ────────────────────────────────────────────────────────────
    mu: float    = 0.3      # Heartbeat rate coefficient
    nu: float    = 0.3      # Blood flow coefficient
    xi: float    = 0.3      # Metabolic drain constant

    # ── EQ_DIGE_05 ────────────────────────────────────────────────────────────
    pi_: float   = 0.1      # Nutrient intake multiplier
    rho_: float  = 0.1      # Activity load factor
    sigma: float = 0.1      # Recovery/drain balance

    # ── EQ_SENS_06 ────────────────────────────────────────────────────────────
    tau: float   = 0.3      # Sensory integration coefficient
    upsilon: float = 0.3    # Attention weight multiplier

    # ── EQ_EMOT_07 ────────────────────────────────────────────────────────────
    phi: float   = 0.1      # Reward/punishment balance
    chi: float   = 0.1      # Neurotransmitter weight

    # ── EQ_PRED_08 ────────────────────────────────────────────────────────────
    psi: float   = 0.05     # Prediction learning rate
    omega: float = 0.05     # Error correction rate

    # ── EQ_ATTN_09 ────────────────────────────────────────────────────────────
    alpha1: float = 0.1     # Stimulus intensity gain
    beta1: float  = 0.1     # Distraction rate

    # ── EQ_WRLD_11 ────────────────────────────────────────────────────────────
    alpha2: float = 0.05    # Resource change coefficient
    beta2: float  = 0.05    # Consumption rate coefficient
    gamma2: float = 0.05    # Social interaction coefficient
    delta2: float = 0.05    # NPC learning rate coefficient

    # ── EQ_LANG_12 ────────────────────────────────────────────────────────────
    lam_lang: float = 0.03  # Language exposure multiplier
    mu_lang: float  = 0.03  # Reinforcement/context multiplier

    # ── Phase 3 ───────────────────────────────────────────────────────────────
    gamma_td: float  = 0.95  # TD discount factor
    beta_forget: float = 0.5 # Memory forgetting exponent
    lam_decay: float = 0.001 # Object permanence decay
    lam_react: float = 0.1   # Object permanence reactivation
    kappa_pred: float = 0.3  # Predictive coding update rate

    # ── Phase 7 ───────────────────────────────────────────────────────────────
    alpha_c: float = 0.4    # Creative synthesis memory blend
    beta_c: float  = 0.4    # Creative synthesis skill-prediction
    gamma_c: float = 0.2    # Creative synthesis stochastic noise
    eta_w: float   = 0.01   # Wisdom learning rate
    lam1: float    = 0.4    # Wisdom consequence weight
    lam2: float    = 0.3    # Wisdom ethical weight
    lam3: float    = 0.2    # Wisdom empathy weight
    lam4: float    = 0.1    # Wisdom impulse suppression
    eta_car: float = 0.02   # Career specialisation rate
    eta_l: float   = 0.005  # Legacy accumulation rate
    entropy_f: float = 0.001 # Legacy forgetting constant

    # ── Cultural / Identity ───────────────────────────────────────────────────
    eta_ce: float  = 0.02   # Cultural embedding rate
    eta_p: float   = 0.001  # Personality stability rate
    eta_s: float   = 0.01   # Skill mastery rate


def load_params(xlsx_path: Path = None) -> A7DOParams:
    """Load parameters from workbook ⚙️ Parameters sheet, fall back to defaults."""
    p = A7DOParams()
    if xlsx_path is None or not xlsx_path.exists():
        return p
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "⚙️ Parameters" not in wb.sheetnames:
            wb.close()
            return p
        ws = wb["⚙️ Parameters"]
        # Map symbol → param attribute
        symbol_map = {
            "α": "alpha", "β": "beta", "γ": "gamma",
            "δ": "delta", "ε": "epsilon", "ζ": "zeta", "η": "eta",
            "θ": "theta", "ι": "iota", "κ": "kappa", "λ": "lam",
            "μ": "mu", "ν": "nu", "ξ": "xi",
            "π_": "pi_", "ρ_": "rho_", "σ": "sigma",
            "τ": "tau", "υ": "upsilon",
            "φ": "phi", "χ": "chi",
            "ψ": "psi", "ω": "omega",
            "α1": "alpha1", "β1": "beta1",
            "α2": "alpha2", "β2": "beta2", "γ2": "gamma2", "δ2": "delta2",
            "λ_lang": "lam_lang", "μ_lang": "mu_lang",
            "γ_td": "gamma_td", "β_forget": "beta_forget",
            "λ_decay": "lam_decay", "λ_react": "lam_react",
            "κ_pred": "kappa_pred",
            "α_c": "alpha_c", "β_c": "beta_c", "γ_c": "gamma_c",
            "η_w": "eta_w", "λ1": "lam1", "λ2": "lam2",
            "λ3": "lam3", "λ4": "lam4",
            "η_car": "eta_car", "η_l": "eta_l", "Entropy_f": "entropy_f",
        }
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[2] is not None:
                sym = str(row[0]).strip()
                if sym in symbol_map:
                    try:
                        setattr(p, symbol_map[sym], float(row[2]))
                    except (TypeError, ValueError):
                        pass
        wb.close()
    except Exception:
        pass
    return p